
from django.http import HttpResponse
from django.contrib import messages
from django.shortcuts import render, redirect, get_object_or_404
from .models import Vehiculo, Tarea, TareaEliminada, ImagenTarea, TareaFinalizada, EntregaVehiculo
from .forms import TareaForm
from .forms import VehiculoForm, EntregaVehiculoForm
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import Group
from django.contrib.auth.models import User, Group
from django.contrib.auth.decorators import user_passes_test
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from datetime import datetime
from reportlab.platypus import (SimpleDocTemplate,Table,TableStyle,Paragraph,Spacer,Image)
from .models import Vehiculo, VehiculoEliminado
from django.db.models import Q
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import enums
from reportlab.lib.units import cm
from django.conf import settings
import os
from reportlab.lib.units import cm
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from django.utils import timezone
from datetime import timedelta




#*********************
#funciones de tecnicos
#*********************

#------------------------
#Vista de inicio tecnicos
#------------------------
@login_required
@user_passes_test(lambda u: u.groups.filter(name='Tecnico').exists())

def tecnicos (request):
    return render(request, 'tecnicos/principal_tecni.html')


#---------------------------------
#Vista lista de vehiculos creados
#----------------------------------

@login_required
@user_passes_test(lambda u: u.groups.filter(name='Tecnico').exists())
def lista_vehiculos(request):

    busqueda = request.GET.get('q', '')

    vehiculos = Vehiculo.objects.all()

    if busqueda:
        vehiculos = vehiculos.filter(
            Q(placa__icontains=busqueda) |
            Q(propietario__icontains=busqueda) |
            Q(marca__icontains=busqueda) |
            Q(color__icontains=busqueda) |
            Q(telefono__icontains=busqueda) |
            Q(Correo__icontains=busqueda)
        )

    return render(
        request,
        'vehiculos/lista.html',
        {
            'vehiculos': vehiculos,
            'busqueda': busqueda
        }
    )
    
#--------------------------
#Funcion de crear cliente
#--------------------------
    
@login_required
@user_passes_test(lambda u: u.groups.filter(name='Tecnico').exists())
def crear_vehiculo(request):
    form = VehiculoForm(request.POST or None,  request.FILES or None)
    if form.is_valid():
        form.save()
        return redirect('lista')
    print(form.errors)  # Para depurar
    return render(request, 'vehiculos/form.html', {'form': form, 'es_edicion':False})

    
#--------------------------
#Funcion de editar cliente
#--------------------------
@login_required
@user_passes_test(lambda u: u.groups.filter(name='Tecnico').exists())

def editar_vehiculo(request, placa):
    vehiculo = get_object_or_404(Vehiculo, pk=placa)
    form = VehiculoForm(request.POST or None, request.FILES or None,
        instance=vehiculo)
    if form.is_valid():
        form.save()
        return redirect('lista')
    return render(request, 'vehiculos/form.html', {'form': form, 'es_edicion':True})

    
#--------------------------
#Funcion de eliminar cliente
#--------------------------

@login_required
@user_passes_test(lambda u: u.groups.filter(name='Tecnico').exists())
def eliminar_vehiculo(request, placa):

    vehiculo = get_object_or_404(Vehiculo, pk=placa)

    if request.method == 'POST':

        motivo = request.POST.get('motivo', '')

        # Guardar registro de eliminación
        VehiculoEliminado.objects.create(
            placa=vehiculo.placa,
            motivo=motivo,
            usuario=request.user
        )

        # Eliminar vehículo
        vehiculo.delete()

        messages.success(
            request,
            f'El vehículo {placa} fue eliminado correctamente.'
        )

        return redirect('lista')

    return redirect('lista')


#-----------------------
#Entregar vehiculo
#------------------------

@login_required
def entregar_vehiculo(request, placa):

    vehiculo = get_object_or_404(
        Vehiculo,
        placa=placa
    )

    if request.method == 'POST':
        
        print("FIRMA RECIBIDA:")
        print(request.POST.get('firma_recibe'))

        form = EntregaVehiculoForm(
            request.POST,
            request.FILES
        )
        if form.is_valid():

            entrega = form.save(commit=False)

            entrega.vehiculo = vehiculo
            entrega.tecnico = request.user
            entrega.firma_recibe = request.POST.get(
                    'firma_recibe',
                    ''
                )

            entrega.save()

            vehiculo.estado = 'Entregado'
            vehiculo.save()

            return redirect('lista')

        else:

            print("ERRORES DEL FORMULARIO:")
            print(form.errors)
            
    else:

        form = EntregaVehiculoForm()

    return render(
        request,
        'tecnicos/entregar_vehiculo.html',
        {
            'vehiculo': vehiculo,
            'form': form
        }
    )


#*********************
# funciones de tareas
#*********************

#--------------------------
# funciones lista de tareas
#--------------------------
@login_required
@user_passes_test(lambda u: u.groups.filter(name='Tecnico').exists())
def lista_tareas(request):

    busqueda = request.GET.get('q', '')

    limite = timezone.now() - timedelta(days=15)

    tareas = Tarea.objects.select_related(
        'vehiculo',
        'tecnico'
    ).exclude(
        estado='Finalizada',
        fecha_finalizacion__lt=limite
    )

    if busqueda:
        tareas = tareas.filter(
            Q(vehiculo__placa__icontains=busqueda) |
            Q(descripcion__icontains=busqueda) |
            Q(estado__icontains=busqueda) |
            Q(tecnico__username__icontains=busqueda)
        )

    return render(
        request,
        'tecnicos/tareas.html',
        {
            'tareas': tareas,
            'busqueda': busqueda
        }
    )
    
#---------------------
# funciones de crear tareas
#---------------------
@login_required
@user_passes_test(lambda u: u.groups.filter(name='Tecnico').exists())
def crear_tarea(request, placa):

    vehiculo = get_object_or_404(
        Vehiculo,
        placa=placa
    )
    
    if vehiculo.estado_actual == 'Entregado':
        messages.error(
            request,
            'No se pueden crear tareas para vehículos entregados.'
        )
        return redirect('lista')

    if request.method == 'POST':
        form = TareaForm(request.POST)

        if form.is_valid():

            tarea = form.save(commit=False)

            tarea.vehiculo = vehiculo
            tarea.tecnico = request.user

            tarea.save()

            return redirect('lista_tareas')

    else:
        form = TareaForm()
        if 'tecnico' in form.fields:
            form.fields['tecnico'].initial = request.user

    return render(
        request,
        'tecnicos/form_tarea.html',
        {'form': form, 'vehiculo':vehiculo,
        'editar':False}
    )

    
#--------------------------
#Funcion de editar tarea
#--------------------------

@login_required
@user_passes_test(lambda u: u.groups.filter(name='Tecnico').exists())
def editar_tarea(request, id):
    
    

    tarea = get_object_or_404(Tarea, id=id)
    
    if tarea.vehiculo.estado_actual == 'Entregado':
        messages.error(
            request,
            'No se pueden editar tareas de vehículos entregados.'
        )
        return redirect('lista_tareas')

    if request.method == 'POST':

        form = TareaForm(
            request.POST,
            instance=tarea
        )

        archivos = request.FILES.getlist('imagenes')

        if form.is_valid():

            tarea_editada = form.save(commit=False)

            # Si se finaliza por primera vez
            if (
                tarea_editada.estado == 'Finalizada'
                and not tarea.fecha_finalizacion
            ):

                tarea_editada.fecha_finalizacion = timezone.now()

                tarea_editada.observaciones_finalizacion = (
                    request.POST.get(
                        'observaciones_finalizacion'
                    )
                )

            tarea_editada.save()

            # Crear historial una sola vez
            if (
                tarea_editada.estado == 'Finalizada'
                and not TareaFinalizada.objects.filter(
                    tarea=tarea_editada
                ).exists()
            ):

                TareaFinalizada.objects.create(
                    tarea=tarea_editada,
                    vehiculo=tarea_editada.vehiculo.placa,
                    descripcion=tarea_editada.descripcion,
                    tecnico=tarea_editada.tecnico.username,
                    fecha_creacion=tarea_editada.fecha_creacion,
                    fecha_finalizacion=tarea_editada.fecha_finalizacion,
                    observaciones=tarea_editada.observaciones_finalizacion
                )

            # Guardar imágenes
            for archivo in archivos:

                ImagenTarea.objects.create(
                    tarea=tarea_editada,
                    imagen=archivo
                )

            return redirect('lista_tareas')

    else:

        form = TareaForm(
            instance=tarea
        )

    return render(
        request,
        'tecnicos/form_tarea.html',
        {
            'form': form,
            'tarea': tarea,
            'vehiculo': tarea.vehiculo,
            'editar': True
        }
    )
    
#----------------------------    
#Lista de tareas finalizadas
#----------------------------

@login_required
@user_passes_test(lambda u: u.groups.filter(name='Administrativo').exists())
def lista_tareas_finalizadas(request):

    tareas = TareaFinalizada.objects.all()\
        .order_by('-fecha_finalizacion')

    return render(
        request,
        'administrativo/tareas_finalizadas.html',
        {'tareas': tareas}
    )    
    
#----------------------------    
#Funcion eliminar tareas 
#----------------------------




@login_required
@user_passes_test(lambda u: u.groups.filter(name='Tecnico').exists())
def eliminar_tarea(request, id):

    tarea = get_object_or_404(Tarea, id=id)

    # No permitir eliminar tareas finalizadas
    if tarea.estado == 'Finalizada':
        messages.error(
            request,
            'Las tareas finalizadas no pueden eliminarse.'
        )
        return redirect('lista_tareas')

    if request.method == 'POST':

        motivo = request.POST.get('motivo')

        TareaEliminada.objects.create(
            vehiculo=tarea.vehiculo.placa,
            descripcion=tarea.descripcion,
            estado=tarea.estado,
            motivo=motivo,
            usuario=request.user
        )

        tarea.delete()

        messages.success(
            request,
            'Tarea eliminada correctamente.'
        )

        return redirect('lista_tareas')

    return redirect('lista_tareas')



#**************************    
#Funciones administrativos
#**************************

#-----------------------------
#vista de inicio panel administrativo
#-----------------------------
@login_required
@user_passes_test(lambda u: u.groups.filter(name='Administrativo').exists())

def administrativos(request):
    return render(request, 'administrativo/principal_admin.html')

#vista del generar informe PDF
@login_required
@user_passes_test(lambda u: u.groups.filter(name='Administrativo').exists())
def informe_view(request):

    return render(
        request,
        'administrativo/informe.html'
    )




#Función de generar informe PDF

@login_required
@user_passes_test(lambda u: u.groups.filter(name='Administrativo').exists())
def informe_tareas_pdf(request):

    placa = request.GET.get('placa')
    cliente = request.GET.get('cliente')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')

    tareas = Tarea.objects.filter(
        estado='Finalizada'
    ).select_related('vehiculo')

    if placa:
        tareas = tareas.filter(
            vehiculo__placa__icontains=placa
        )

    if cliente:
        tareas = tareas.filter(
            vehiculo__propietario__icontains=cliente
        )

    if fecha_inicio:
        tareas = tareas.filter(
            fecha_creacion__gte=fecha_inicio
        )

    if fecha_fin:
        tareas = tareas.filter(
            fecha_creacion__lte=fecha_fin
        )

    response = HttpResponse(
        content_type='application/pdf'
    )

    response['Content-Disposition'] = (
        'attachment; filename="Informe_Tareas.pdf"'
    )

    doc = SimpleDocTemplate(response)

    elementos = []

    estilos = getSampleStyleSheet()

    # LOGO
    logo_path = os.path.join(
        settings.BASE_DIR,
        'static',
        'img',
        'logo.png'
    )

    if os.path.exists(logo_path):

        logo = Image(
            logo_path,
            width=3*cm,
            height=3*cm
        )

        elementos.append(logo)

    # TITULO
    titulo = Paragraph(
        "<b>GRUPO ELITE</b><br/>"
        "INFORME DE TAREAS FINALIZADAS",
        estilos['Title']
    )

    elementos.append(titulo)
    elementos.append(Spacer(1, 12))

    # FILTROS
    filtros = f"""
    <b>Filtros Aplicados</b><br/>
    Placa: {placa if placa else 'Todas'}<br/>
    Cliente: {cliente if cliente else 'Todos'}<br/>
    Fecha Inicial: {fecha_inicio if fecha_inicio else 'Sin filtro'}<br/>
    Fecha Final: {fecha_fin if fecha_fin else 'Sin filtro'}<br/>
    Total Registros: {tareas.count()}
    """

    elementos.append(
        Paragraph(
            filtros,
            estilos['Normal']
        )
    )

    elementos.append(Spacer(1, 15))

    # TABLA
    datos = [[
        'Placa',
        'Cliente',
        'Técnico',
        'Estado',
        'Fecha'
    ]]

    for tarea in tareas:

        datos.append([
            tarea.vehiculo.placa,
            tarea.vehiculo.propietario,
            tarea.tecnico.username if tarea.tecnico else "No asignado",
            tarea.estado,
            str(tarea.fecha_creacion)
        ])

    tabla = Table(
        datos,
        colWidths=[
            80,
            120,
            100,
            80,
            80
        ]
    )

    tabla.setStyle(TableStyle([

        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#003366')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),

        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),

        ('GRID', (0,0), (-1,-1), 1, colors.black),

        ('ALIGN', (0,0), (-1,-1), 'CENTER'),

        ('BACKGROUND', (0,1), (-1,-1), colors.whitesmoke)

    ]))

    elementos.append(tabla)

    elementos.append(Spacer(1, 20))

    elementos.append(
        Paragraph(
            f"Fecha de generación: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
            estilos['Normal']
        )
    )
    
    doc.build(
    elementos,
    onFirstPage=agregar_pie_pagina,
    onLaterPages=agregar_pie_pagina
)
    return response


def agregar_pie_pagina(canvas, doc):

    canvas.saveState()

    canvas.setFont(
            'Helvetica',
            9
        )

    pagina = canvas.getPageNumber()

    texto = f"Página {pagina}"

    canvas.drawRightString(
            19 * cm,
            1 * cm,
            texto
        )
    canvas.restoreState()

#Vista de Excel


def informe_tareas_excel(request):

    placa = request.GET.get('placa')
    cliente = request.GET.get('cliente')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')

    tareas = Tarea.objects.filter(
        estado='Finalizada'
    ).select_related('vehiculo')

    if placa:
        tareas = tareas.filter(
            vehiculo__placa__icontains=placa
        )

    if cliente:
        tareas = tareas.filter(
            vehiculo__propietario__icontains=cliente
        )

    if fecha_inicio:
        tareas = tareas.filter(
            fecha_creacion__gte=fecha_inicio
        )

    if fecha_fin:
        tareas = tareas.filter(
            fecha_creacion__lte=fecha_fin
        )

    wb = Workbook()
    ws = wb.active

    ws.title = "Informe Tareas"

    # Título
    ws.merge_cells('A1:F1')
    ws['A1'] = "GRUPO ELITE - INFORME DE TAREAS FINALIZADAS"

    ws['A1'].font = Font(
        bold=True,
        size=14
    )

    ws['A1'].alignment = Alignment(
        horizontal='center'
    )

    # Encabezados
    encabezados = [
        'Placa',
        'Cliente',
        'Técnico',
        'Descripción',
        'Estado',
        'Fecha'
    ]

    fila = 3

    for col_num, encabezado in enumerate(encabezados, 1):

        celda = ws.cell(
            row=fila,
            column=col_num
        )

        celda.value = encabezado

        celda.font = Font(
            bold=True,
            color='FFFFFF'
        )

        celda.fill = PatternFill(
            'solid',
            fgColor='1F4E78'
        )

    # Datos
    fila += 1

    for tarea in tareas:

        ws.cell(fila, 1, tarea.vehiculo.placa)
        ws.cell(fila, 2, tarea.vehiculo.propietario)

        ws.cell(
            fila,
            3,
            tarea.tecnico.username if tarea.tecnico else 'No asignado'
        )

        ws.cell(
            fila,
            4,
            tarea.descripcion
        )

        ws.cell(
            fila,
            5,
            tarea.estado
        )

        ws.cell(
            fila,
            6,
            str(tarea.fecha_creacion)
        )

        fila += 1

    # Ajustar ancho columnas
    for columna in ws.columns:

        longitud = max(
            len(str(celda.value))
            if celda.value else 0
            for celda in columna
        )

        ws.column_dimensions[
            columna[0].column_letter
        ].width = longitud + 5

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    response[
        'Content-Disposition'
    ] = 'attachment; filename=Informe_Tareas.xlsx'

    wb.save(response)

    return response


#Vista de vehiculos eliminados


@login_required
@user_passes_test(lambda u: u.groups.filter(name='Administrativo').exists())
def historial_eliminados(request):

    historial = VehiculoEliminado.objects.select_related(
        'usuario'
    ).order_by('-fecha')

    return render(
        request,
        'administrativo/historial_eliminados.html',
        {'historial': historial}
    )


#lista clientes en el panel administrativos

from django.contrib.auth.decorators import login_required, user_passes_test



@login_required
@user_passes_test(lambda u: u.groups.filter(name='Administrativo').exists())
def lista_vehiculos_admin(request):

    busqueda = request.GET.get('q', '')

    vehiculos = Vehiculo.objects.all()

    if busqueda:
        vehiculos = vehiculos.filter(
            Q(placa__icontains=busqueda) |
            Q(propietario__icontains=busqueda) |
            Q(marca__icontains=busqueda) |
            Q(color__icontains=busqueda) |
            Q(telefono__icontains=busqueda) |
            Q(Correo__icontains=busqueda)
        )

    return render(
        request,
        'administrativo/lista_clientes.html',
        {
            'vehiculos': vehiculos,
            'busqueda': busqueda
        }
    )

#lista de tareas eliminadas

@login_required
@user_passes_test(lambda u: u.groups.filter(name='Administrativo').exists())
def historial_tareas_eliminadas(request):

    tareas = TareaEliminada.objects.select_related(
        'usuario'
    ).order_by('-fecha')

    return render(
        request,
        'administrativo/historial_tareas_eliminadas.html',
        {'tareas': tareas}
    )
    
# Lista de los estados de las tareas en vistas administrativa

@login_required
@user_passes_test(lambda u: u.groups.filter(name='Administrativo').exists())
def lista_tareas_admin(request):

    tareas = Tarea.objects.select_related(
        'vehiculo',
        'tecnico'
    ).order_by('-fecha_actualizacion')

    return render(
        request,
        'administrativo/lista_tareas_admin.html',
        {'tareas': tareas}
    )
    
    
@login_required
@user_passes_test(lambda u: u.groups.filter(name__in=['Tecnico', 'Administrativo']).exists())

def imagenes_tarea(request, tarea_id):

    tarea = get_object_or_404(Tarea, id=tarea_id)

    imagenes = tarea.imagenes.all().order_by('-fecha_subida')

    return render(
        request,
        'tecnicos/imagenes_tarea.html',
        {
            'tarea': tarea,
            'imagenes': imagenes
        }
    )
    
    
def imagenes_tarea_finalizada(request, id):
    tarea_finalizada = TareaFinalizada.objects.get(id=id)

    imagenes = tarea_finalizada.tarea.imagenes.all()

    return render(request, 'administrativo/imagenes_tarea_admin.html', {
        'tarea': tarea_finalizada.tarea,
        'imagenes': imagenes
    })
